"""
Tự động lấy top 100 ví nắm giữ (holders) của các token trên TON qua TonAPI,
rồi cập nhật vào file Excel theo đúng format đang dùng:
- Mỗi sheet = 1 token
- Mỗi dòng = 1 địa chỉ ví
- Mỗi lần chạy = thêm 1 cột ngày giờ + 1 cột "Chênh lệch" (so với lần chạy trước)
  Chênh lệch dương -> tô xanh, âm -> tô đỏ
- Ví mới xuất hiện trong top 100 -> thêm dòng mới
- Ví cũ rớt khỏi top 100 -> coi số dư ngày đó là 0

Chạy thủ công:
    export TONAPI_KEY=xxxx   # (tùy chọn, không có vẫn chạy được nhưng dễ bị rate-limit)
    python update_tracker.py

Cấu hình token cần theo dõi ở biến TOKENS bên dưới.
"""

import os
import time
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timezone, timedelta

# ============ CẤU HÌNH ============

# Mỗi token: tên hiển thị (cũng là tên sheet) -> địa chỉ contract (jetton master) trên TON
TOKENS = {
    "DIDI": "EQCRUitj7ehYvSzZKTyhq02-HpbhLNgAvnMF5I7Dx31QxIAH",
    "YODA": "EQC7vuKEYLdC72YhUWt3AUVA-Oi66Q1DxTHXH7r6pXaV50j7",
}

TOP_N = 100  # số ví top đầu muốn theo dõi mỗi token
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "theo_doi_vi.xlsx")
TONAPI_BASE = "https://tonapi.io"
TONAPI_KEY = os.environ.get("TONAPI_KEY", "")  # optional nhưng khuyến khích có

# Giờ Việt Nam để đặt tên cột ngày cho dễ đọc
VN_TZ = timezone(timedelta(hours=7))

# ============ STYLE DÙNG CHUNG ============

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NORMAL_FONT = Font(name="Arial")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT = Font(name="Arial", color="006100")
RED_FONT = Font(name="Arial", color="9C0006")


def tonapi_get(path, params=None):
    headers = {}
    if TONAPI_KEY:
        headers["Authorization"] = f"Bearer {TONAPI_KEY}"
    for attempt in range(5):
        resp = requests.get(f"{TONAPI_BASE}{path}", params=params, headers=headers, timeout=30)
        if resp.status_code == 429:
            # rate limited -> chờ rồi thử lại
            time.sleep(3 * (attempt + 1))
            continue
        resp.raise_for_status()
        return resp.json()
    resp.raise_for_status()


def get_jetton_decimals(jetton_address):
    info = tonapi_get(f"/v2/jettons/{jetton_address}")
    meta = info.get("metadata", {})
    return int(meta.get("decimals", 9)), meta.get("symbol", "")


def get_top_holders(jetton_address, limit=100):
    data = tonapi_get(f"/v2/jettons/{jetton_address}/holders", params={"limit": limit, "offset": 0})
    holders = []
    for item in data.get("addresses", []):
        owner = item.get("owner", {}) or {}
        addr = owner.get("address") or item.get("address")
        balance_raw = item.get("balance", "0")
        holders.append((addr, int(balance_raw)))
    return holders


def load_or_create_workbook(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    # xóa sheet mặc định, mỗi token sẽ tự tạo sheet riêng
    default_sheet = wb.active
    wb.remove(default_sheet)
    return wb


def ensure_sheet(wb, token_name):
    if token_name in wb.sheetnames:
        return wb[token_name]
    ws = wb.create_sheet(token_name)
    ws["A1"] = "Địa chỉ ví"
    ws["B1"] = "Token"
    for col_letter in ("A", "B"):
        c = ws[f"{col_letter}1"]
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 12
    ws.freeze_panes = "C2"
    return ws


def update_sheet(ws, token_name, holders, timestamp_label, decimals):
    # holders: list[(address, balance_raw_int)]
    human_balances = {addr: bal / (10 ** decimals) for addr, bal in holders}

    last_col = ws.max_column
    if last_col >= 3 and ws.cell(row=1, column=last_col).value == "Chênh lệch":
        prev_col = last_col - 1
    elif last_col >= 3:
        prev_col = last_col
    else:
        prev_col = None  # sheet vừa tạo, chưa có cột ngày nào

    existing = {}
    for r in range(2, ws.max_row + 1):
        addr = ws.cell(row=r, column=1).value
        if addr:
            existing[addr] = r

    # ví cũ không còn trong top N hôm nay -> balance = 0
    all_addrs = dict(human_balances)
    for addr in existing:
        if addr not in all_addrs:
            all_addrs[addr] = 0

    next_row = ws.max_row + 1

    new_col = last_col + 1
    diff_col = new_col + 1
    ws.cell(row=1, column=new_col, value=timestamp_label)
    ws.cell(row=1, column=diff_col, value="Chênh lệch")
    for c in (new_col, diff_col):
        hc = ws.cell(row=1, column=c)
        hc.font = HEADER_FONT
        hc.fill = HEADER_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = BORDER
        ws.column_dimensions[hc.column_letter].width = 16

    for addr, bal in all_addrs.items():
        if addr in existing:
            r = existing[addr]
            prev_val = ws.cell(row=r, column=prev_col).value if prev_col else None
        else:
            r = next_row
            next_row += 1
            existing[addr] = r
            ws.cell(row=r, column=1, value=addr).font = NORMAL_FONT
            ws.cell(row=r, column=1).border = BORDER
            ws.cell(row=r, column=2, value=token_name).font = NORMAL_FONT
            ws.cell(row=r, column=2).border = BORDER
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
            prev_val = None

        cell = ws.cell(row=r, column=new_col, value=round(bal, 4))
        cell.font = NORMAL_FONT
        cell.number_format = "#,##0.####"
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right")

        dcell = ws.cell(row=r, column=diff_col)
        dcell.border = BORDER
        dcell.number_format = "+#,##0.####;-#,##0.####;0"
        dcell.alignment = Alignment(horizontal="right")
        if isinstance(prev_val, (int, float)):
            diff = bal - prev_val
            dcell.value = round(diff, 4)
            if diff > 0:
                dcell.fill = GREEN_FILL
                dcell.font = GREEN_FONT
            elif diff < 0:
                dcell.fill = RED_FILL
                dcell.font = RED_FONT
            else:
                dcell.font = NORMAL_FONT
        else:
            dcell.font = NORMAL_FONT


def main():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb = load_or_create_workbook(EXCEL_PATH)

    now_vn = datetime.now(VN_TZ)
    timestamp_label = now_vn.strftime("%d/%m/%Y %H:%M")

    for token_name, jetton_address in TOKENS.items():
        print(f"[{token_name}] fetching decimals + top {TOP_N} holders...")
        decimals, symbol = get_jetton_decimals(jetton_address)
        holders = get_top_holders(jetton_address, limit=TOP_N)
        print(f"[{token_name}] got {len(holders)} holders (decimals={decimals}, symbol={symbol})")

        ws = ensure_sheet(wb, token_name)
        update_sheet(ws, token_name, holders, timestamp_label, decimals)

    wb.save(EXCEL_PATH)
    print(f"Saved -> {EXCEL_PATH}")


if __name__ == "__main__":
    main()
